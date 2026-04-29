# core/dashboard.py
# ============================================================
# Excel Dashboard Generator
# ============================================================
# Reads the enriched CSV output and produces a formatted Excel
# workbook with three sheets:
#
#   1. Summary   — KPI cards, cost breakdowns, and charts
#   2. By Location — cost breakdown per Azure region
#   3. Raw Data  — full filterable table with NDF highlighting
#
# Uses openpyxl exclusively — no dependency on Excel being
# installed. Works on macOS, Linux, and Windows.
#
# Entry point: generate(input_csv, output_xlsx)
# ============================================================

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.console import header, success, warn, info, dim


# ── Palette ────────────────────────────────────────────────
# Central colour definitions — change here to restyle everything.

class Colours:
    # Backgrounds
    DARK_BLUE    = "1F3864"   # Header rows, title bar
    MID_BLUE     = "2E75B6"   # Sub-headers, KPI labels
    LIGHT_BLUE   = "D6E4F0"   # Alternating row tint
    WHITE        = "FFFFFF"
    LIGHT_GREY   = "F2F2F2"   # Alternating rows
    NDF_YELLOW   = "FFF2CC"   # NDF cell highlight
    NDF_BORDER   = "FFD700"   # NDF cell border
    ZERO_GREY    = "EDEDED"   # Zero-cost row background
    GREEN_ACCENT = "375623"   # Positive/total text

    # Font colours
    WHITE_TEXT   = "FFFFFF"
    DARK_TEXT    = "1F1F1F"
    GREY_TEXT    = "595959"
    GREEN_TEXT   = "375623"
    RED_TEXT     = "C00000"


# ── Border helpers ─────────────────────────────────────────

def _thin_border(top=True, bottom=True, left=True, right=True) -> Border:
    thin = Side(style="thin", color=Colours.DARK_TEXT)
    none = Side(style=None)
    return Border(
        top    = thin if top    else none,
        bottom = thin if bottom else none,
        left   = thin if left   else none,
        right  = thin if right  else none,
    )


def _thick_bottom() -> Border:
    return Border(bottom=Side(style="medium", color=Colours.DARK_BLUE))


# ── Cell style helpers ─────────────────────────────────────

def _style_header_cell(cell, text: str, font_size: int = 11) -> None:
    """Dark blue background, white bold text — used for column headers."""
    cell.value     = text
    cell.font      = Font(name="Calibri", bold=True, color=Colours.WHITE_TEXT,
                          size=font_size)
    cell.fill      = PatternFill("solid", fgColor=Colours.DARK_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border    = _thin_border()


def _style_subheader_cell(cell, text: str) -> None:
    """Mid blue background — used for section labels."""
    cell.value     = text
    cell.font      = Font(name="Calibri", bold=True, color=Colours.WHITE_TEXT,
                          size=10)
    cell.fill      = PatternFill("solid", fgColor=Colours.MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _thin_border()


def _style_kpi_label(cell, text: str) -> None:
    """KPI card label row."""
    cell.value     = text
    cell.font      = Font(name="Calibri", bold=True, color=Colours.WHITE_TEXT,
                          size=9)
    cell.fill      = PatternFill("solid", fgColor=Colours.MID_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_kpi_value(cell, value, number_format: str = None) -> None:
    """KPI card value row — large, prominent number."""
    cell.value     = value
    cell.font      = Font(name="Calibri", bold=True,
                          color=Colours.DARK_BLUE, size=16)
    cell.fill      = PatternFill("solid", fgColor=Colours.LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()
    if number_format:
        cell.number_format = number_format


def _style_data_cell(cell, value, row_even: bool = True,
                     number_format: str = None, is_ndf: bool = False) -> None:
    """Standard data cell with alternating row shading."""
    cell.value = value

    if is_ndf:
        # Highlight NDF cells in yellow so they stand out immediately
        cell.fill   = PatternFill("solid", fgColor=Colours.NDF_YELLOW)
        cell.font   = Font(name="Calibri", color="7F6000", italic=True, size=10)
        cell.border = Border(
            top    = Side(style="thin", color=Colours.NDF_BORDER),
            bottom = Side(style="thin", color=Colours.NDF_BORDER),
            left   = Side(style="thin", color=Colours.NDF_BORDER),
            right  = Side(style="thin", color=Colours.NDF_BORDER),
        )
    else:
        bg = Colours.WHITE if row_even else Colours.LIGHT_GREY
        cell.fill   = PatternFill("solid", fgColor=bg)
        cell.font   = Font(name="Calibri", color=Colours.DARK_TEXT, size=10)
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")

    if number_format:
        cell.number_format = number_format


# ── Sheet 1: Summary ───────────────────────────────────────

def _build_summary_sheet(ws, df: pd.DataFrame, source_file: str) -> None:
    """
    Build the Summary sheet.

    Layout:
      Row 1-2   : Title banner
      Row 3-4   : Metadata (date, source file)
      Row 6-8   : KPI cards (total cost, resources, subscriptions, NDF count)
      Row 10+   : Cost by resource type table + bar chart
      Row 10+   : Cost by subscription table + pie chart
    """
    ws.title = "Summary"

    # ── Column widths ──────────────────────────────────────
    col_widths = [2, 22, 16, 16, 16, 16, 2, 22, 16, 2]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title banner ───────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    ws.merge_cells("B1:I1")
    title_cell = ws["B1"]
    title_cell.value     = "FinOps Cost Report — Azure"
    title_cell.font      = Font(name="Calibri", bold=True, size=20,
                                color=Colours.WHITE_TEXT)
    title_cell.fill      = PatternFill("solid", fgColor=Colours.DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B2:I2")
    date_cell = ws["B2"]
    date_cell.value     = (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} "
                           f"  |  Source: {source_file}")
    date_cell.font      = Font(name="Calibri", size=9,
                               color=Colours.WHITE_TEXT, italic=True)
    date_cell.fill      = PatternFill("solid", fgColor=Colours.MID_BLUE)
    date_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── KPI cards ──────────────────────────────────────────
    # Four cards side by side: Total Cost | Resources | Subscriptions | NDF Count
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 36

    costs       = pd.to_numeric(df["estimated_cost_usd"], errors="coerce").fillna(0)
    total_cost  = costs.sum()
    n_resources = len(df)
    n_subs      = df["subscription_name"].nunique()
    ndf_count   = (df == "NDF").sum().sum()  # Count across all columns

    kpis = [
        ("Total Est. Monthly Cost (USD)", f"${total_cost:,.2f}", "$#,##0.00"),
        ("Total Resources",               n_resources,           "#,##0"),
        ("Subscriptions",                 n_subs,                "#,##0"),
        ("NDF Fields",                    int(ndf_count),        "#,##0"),
    ]

    # KPI positions: (label_col, value_col) pairs
    kpi_cols = [("B", "C"), ("D", "E"), ("F", "G"), ("H", "I")]

    for (label, value, fmt), (lcol, vcol) in zip(kpis, kpi_cols):
        ws.merge_cells(f"{lcol}4:{vcol}4")
        ws.merge_cells(f"{lcol}5:{vcol}5")
        _style_kpi_label(ws[f"{lcol}4"], label)
        _style_kpi_value(ws[f"{lcol}5"], value, fmt)

    # ── Cost by resource type ──────────────────────────────
    ws.row_dimensions[7].height = 20
    ws.merge_cells("B7:E7")
    _style_subheader_cell(ws["B7"], "  Cost by Resource Type")

    type_headers = ["Resource Type", "Resources", "Est. Cost (USD)", "% of Total"]
    for col_idx, hdr in enumerate(type_headers, start=2):
        _style_header_cell(ws.cell(row=8, column=col_idx), hdr, font_size=10)

    df_copy        = df.copy()
    df_copy["_cost"] = costs
    by_type        = (df_copy.groupby("resource_type")
                      .agg(resources=("resource_name", "count"),
                           cost=("_cost", "sum"))
                      .sort_values("cost", ascending=False)
                      .reset_index())

    type_data_start = 9
    for i, row in by_type.iterrows():
        r        = type_data_start + i
        even     = i % 2 == 0
        pct      = (row["cost"] / total_cost * 100) if total_cost > 0 else 0
        rtype    = row["resource_type"].split("/")[-1]  # Shorten for readability

        _style_data_cell(ws.cell(r, 2), rtype,           even)
        _style_data_cell(ws.cell(r, 3), int(row["resources"]), even, "#,##0")
        _style_data_cell(ws.cell(r, 4), row["cost"],     even, "$#,##0.00")
        _style_data_cell(ws.cell(r, 5), pct / 100,       even, "0.0%")

    # Total row
    total_r = type_data_start + len(by_type)
    ws.row_dimensions[total_r].height = 16
    for col_idx in range(2, 6):
        c = ws.cell(total_r, col_idx)
        c.fill   = PatternFill("solid", fgColor=Colours.DARK_BLUE)
        c.font   = Font(name="Calibri", bold=True,
                        color=Colours.WHITE_TEXT, size=10)
        c.border = _thick_bottom()
    ws.cell(total_r, 2).value          = "TOTAL"
    ws.cell(total_r, 3).value          = n_resources
    ws.cell(total_r, 3).number_format  = "#,##0"
    ws.cell(total_r, 4).value          = total_cost
    ws.cell(total_r, 4).number_format  = "$#,##0.00"
    ws.cell(total_r, 5).value          = 1.0
    ws.cell(total_r, 5).number_format  = "0.0%"

    # ── Bar chart: cost by resource type ───────────────────
    bar = BarChart()
    bar.type         = "col"
    bar.title        = "Estimated Monthly Cost by Resource Type (USD)"
    bar.y_axis.title = "USD"
    bar.x_axis.title = "Resource Type"
    bar.style        = 10
    bar.width        = 22
    bar.height       = 14
    bar.grouping     = "clustered"

    # Data: cost column (col 4 = D) from type_data_start to last data row
    data_rows = len(by_type)
    data_ref  = Reference(ws,
                          min_col=4, min_row=type_data_start,
                          max_row=type_data_start + data_rows - 1)
    cats_ref  = Reference(ws,
                          min_col=2, min_row=type_data_start,
                          max_row=type_data_start + data_rows - 1)
    bar.add_data(data_ref)
    bar.set_categories(cats_ref)
    bar.series[0].title = None

    # Place the chart to the right of the table
    ws.add_chart(bar, "G7")

    # ── Cost by subscription ───────────────────────────────
    sub_start_row = total_r + 2
    ws.row_dimensions[sub_start_row].height = 20
    ws.merge_cells(f"B{sub_start_row}:E{sub_start_row}")
    _style_subheader_cell(ws[f"B{sub_start_row}"], "  Cost by Subscription")

    sub_headers = ["Subscription", "Resources", "Est. Cost (USD)", "% of Total"]
    for col_idx, hdr in enumerate(sub_headers, start=2):
        _style_header_cell(ws.cell(sub_start_row + 1, col_idx), hdr, font_size=10)

    by_sub = (df_copy.groupby("subscription_name")
              .agg(resources=("resource_name", "count"),
                   cost=("_cost", "sum"))
              .sort_values("cost", ascending=False)
              .reset_index())

    sub_data_start = sub_start_row + 2
    for i, row in by_sub.iterrows():
        r    = sub_data_start + i
        even = i % 2 == 0
        pct  = (row["cost"] / total_cost * 100) if total_cost > 0 else 0

        _style_data_cell(ws.cell(r, 2), row["subscription_name"], even)
        _style_data_cell(ws.cell(r, 3), int(row["resources"]),    even, "#,##0")
        _style_data_cell(ws.cell(r, 4), row["cost"],              even, "$#,##0.00")
        _style_data_cell(ws.cell(r, 5), pct / 100,                even, "0.0%")

    # Total row for subscriptions
    sub_total_r = sub_data_start + len(by_sub)
    for col_idx in range(2, 6):
        c = ws.cell(sub_total_r, col_idx)
        c.fill   = PatternFill("solid", fgColor=Colours.DARK_BLUE)
        c.font   = Font(name="Calibri", bold=True,
                        color=Colours.WHITE_TEXT, size=10)
        c.border = _thick_bottom()
    ws.cell(sub_total_r, 2).value         = "TOTAL"
    ws.cell(sub_total_r, 3).value         = n_resources
    ws.cell(sub_total_r, 3).number_format = "#,##0"
    ws.cell(sub_total_r, 4).value         = total_cost
    ws.cell(sub_total_r, 4).number_format = "$#,##0.00"
    ws.cell(sub_total_r, 5).value         = 1.0
    ws.cell(sub_total_r, 5).number_format = "0.0%"

    # ── Pie chart: cost by subscription ───────────────────
    pie = PieChart()
    pie.title  = "Cost Share by Subscription"
    pie.style  = 10
    pie.width  = 16
    pie.height = 12

    pie_data = Reference(ws,
                         min_col=4, min_row=sub_data_start,
                         max_row=sub_data_start + len(by_sub) - 1)
    pie_cats = Reference(ws,
                         min_col=2, min_row=sub_data_start,
                         max_row=sub_data_start + len(by_sub) - 1)
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)

    ws.add_chart(pie, f"G{sub_start_row}")


# ── Sheet 2: By Location ───────────────────────────────────

def _build_location_sheet(ws, df: pd.DataFrame) -> None:
    """
    Build the By Location sheet — cost breakdown per Azure region.
    """
    ws.title = "By Location"

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    # Title
    ws.row_dimensions[1].height = 26
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value     = "Cost Breakdown by Azure Region"
    c.font      = Font(name="Calibri", bold=True, size=14,
                       color=Colours.WHITE_TEXT)
    c.fill      = PatternFill("solid", fgColor=Colours.DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["Location", "Resources", "Est. Cost (USD)", "% of Total"]
    for col_idx, hdr in enumerate(headers, start=2):
        _style_header_cell(ws.cell(3, col_idx), hdr, font_size=10)

    # Data
    costs = pd.to_numeric(df["estimated_cost_usd"], errors="coerce").fillna(0)
    df_copy         = df.copy()
    df_copy["_cost"] = costs
    total_cost      = costs.sum()

    by_loc = (df_copy.groupby("location")
              .agg(resources=("resource_name", "count"),
                   cost=("_cost", "sum"))
              .sort_values("cost", ascending=False)
              .reset_index())

    data_start = 4
    for i, row in by_loc.iterrows():
        r    = data_start + i
        even = i % 2 == 0
        pct  = (row["cost"] / total_cost * 100) if total_cost > 0 else 0

        _style_data_cell(ws.cell(r, 2), row["location"],        even)
        _style_data_cell(ws.cell(r, 3), int(row["resources"]), even, "#,##0")
        _style_data_cell(ws.cell(r, 4), row["cost"],           even, "$#,##0.00")
        _style_data_cell(ws.cell(r, 5), pct / 100,             even, "0.0%")

    # Total row
    total_r = data_start + len(by_loc)
    for col_idx in range(2, 6):
        c = ws.cell(total_r, col_idx)
        c.fill   = PatternFill("solid", fgColor=Colours.DARK_BLUE)
        c.font   = Font(name="Calibri", bold=True,
                        color=Colours.WHITE_TEXT, size=10)
    ws.cell(total_r, 2).value         = "TOTAL"
    ws.cell(total_r, 3).value         = len(df)
    ws.cell(total_r, 3).number_format = "#,##0"
    ws.cell(total_r, 4).value         = total_cost
    ws.cell(total_r, 4).number_format = "$#,##0.00"
    ws.cell(total_r, 5).value         = 1.0
    ws.cell(total_r, 5).number_format = "0.0%"

    # Bar chart: cost by location
    bar = BarChart()
    bar.type         = "bar"   # Horizontal bars — better for location names
    bar.title        = "Estimated Monthly Cost by Region (USD)"
    bar.y_axis.title = "Region"
    bar.x_axis.title = "USD"
    bar.style        = 10
    bar.width        = 20
    bar.height       = 14

    data_ref = Reference(ws, min_col=4, min_row=data_start,
                         max_row=data_start + len(by_loc) - 1)
    cats_ref = Reference(ws, min_col=2, min_row=data_start,
                         max_row=data_start + len(by_loc) - 1)
    bar.add_data(data_ref)
    bar.set_categories(cats_ref)
    bar.series[0].title = None

    ws.add_chart(bar, "G3")


# ── Sheet 3: Raw Data ──────────────────────────────────────

def _build_raw_data_sheet(ws, df: pd.DataFrame) -> None:
    """
    Build the Raw Data sheet.

    Features:
    - Frozen header row
    - Auto-filter on all columns
    - NDF cells highlighted in yellow
    - Zero-cost rows shaded in light grey
    - Formatted as a named Excel Table for easy slicing
    """
    ws.title = "Raw Data"

    # Column definitions: (header text, data key, width, number_format)
    columns = [
        ("Subscription ID",        "subscription_id",    22, None),
        ("Subscription Name",      "subscription_name",  20, None),
        ("Resource Group",         "resource_group",     22, None),
        ("Resource Name",          "resource_name",      28, None),
        ("Resource Type",          "resource_type",      38, None),
        ("Location",               "location",           14, None),
        ("SKU",                    "sku",                14, None),
        ("Size",                   "size",               18, None),
        ("Unit",                   "unit",               24, None),
        ("Quantity",               "quantity",           10, "#,##0.00"),
        ("Unit Price (USD)",       "unit_price_usd",     14, "$#,##0.0000"),
        ("Est. Monthly Cost (USD)","estimated_cost_usd", 18, "$#,##0.00"),
    ]

    # Set column widths and header row
    ws.row_dimensions[1].height = 22
    for col_idx, (hdr, _, width, _fmt) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        _style_header_cell(ws.cell(1, col_idx), hdr, font_size=10)

    # Data rows
    costs = pd.to_numeric(df["estimated_cost_usd"], errors="coerce").fillna(0)

    for row_idx, (df_i, row) in enumerate(df.iterrows(), start=2):
        even      = row_idx % 2 == 0
        zero_cost = costs.iloc[row_idx - 2] == 0

        ws.row_dimensions[row_idx].height = 16

        for col_idx, (_, key, _, fmt) in enumerate(columns, start=1):
            value  = row.get(key, "NDF")
            is_ndf = str(value).strip().upper() == "NDF"

            # Convert numeric strings back to floats for numeric columns
            if fmt and not is_ndf:
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass

            # Zero-cost rows: light grey background (overrides alternating)
            # NDF cells: yellow (handled inside _style_data_cell)
            if zero_cost and not is_ndf:
                c = ws.cell(row_idx, col_idx)
                c.value          = value
                c.fill           = PatternFill("solid", fgColor=Colours.ZERO_GREY)
                c.font           = Font(name="Calibri", color=Colours.GREY_TEXT,
                                        size=10, italic=True)
                c.border         = _thin_border()
                c.alignment      = Alignment(vertical="center")
                if fmt:
                    c.number_format = fmt
            else:
                _style_data_cell(ws.cell(row_idx, col_idx),
                                 value, even, fmt, is_ndf)

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    # Add Excel auto-filter to all columns
    last_col = get_column_letter(len(columns))
    last_row = 1 + len(df)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Wrap the data in an Excel Table for easy filtering/sorting in Excel
    table = Table(
        displayName="FinOpsData",
        ref=f"A1:{last_col}{last_row}",
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


# ── Public entry point ─────────────────────────────────────

def generate(input_csv: str, output_xlsx: str) -> None:
    """
    Read an enriched FinOps CSV and write a formatted Excel dashboard.

    Args:
        input_csv:   Path to the CSV produced by export_csv() or
                     the sample data file.
        output_xlsx: Path for the output Excel file.

    Raises:
        SystemExit: If the CSV cannot be read or the file cannot be written.
    """
    header(f"[DASHBOARD] Reading data from: {input_csv}")

    try:
        # Read all columns as strings first, then convert numerics explicitly.
        # This prevents pandas from misinterpreting IDs as numbers.
        df = pd.read_csv(input_csv, dtype=str, encoding="utf-8-sig")
    except FileNotFoundError:
        from core.console import error
        error(f"[DASHBOARD] File not found: {input_csv}")
        raise SystemExit(1)

    if df.empty:
        from core.console import warn as cwarn
        cwarn("[DASHBOARD] CSV is empty — no dashboard generated.")
        return

    info(f"[DASHBOARD] Loaded {len(df)} rows.")
    info(f"[DASHBOARD] Building workbook...")

    wb = Workbook()

    # Remove the default empty sheet that openpyxl creates
    wb.remove(wb.active)

    # Build each sheet
    dim("  → Summary sheet")
    _build_summary_sheet(wb.create_sheet(), df, input_csv)

    dim("  → By Location sheet")
    _build_location_sheet(wb.create_sheet(), df)

    dim("  → Raw Data sheet")
    _build_raw_data_sheet(wb.create_sheet(), df)

    # Save
    try:
        wb.save(output_xlsx)
        success(f"[DASHBOARD] Saved → {output_xlsx}")
    except IOError as e:
        from core.console import error as cerr
        cerr(f"[DASHBOARD] Failed to save: {e}")
        raise SystemExit(1)
